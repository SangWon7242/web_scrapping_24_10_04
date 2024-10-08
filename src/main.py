import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 1. 웹 페이지 요청
url = 'https://news.naver.com/section/104'  # 스크래핑할 URL
response = requests.get(url)

# 2. 요청이 성공했는지 확인

if response.status_code == 200:
  print("웹 페이지를 성공적으로 가져왔습니다!")    

# BeautifulSoup 객체 생성
# response.text는 HTML 소스 코드 텍스트임
html_data = BeautifulSoup(response.text, 'html.parser')  

# print(html_data.select('.sa_text_strong'))

headline_news_list = html_data.select('.sa_text_strong')
news_title_list = []

# 1차 가공
for news_title in headline_news_list:
  # print(news_title.get_text())
  news_title_txt = news_title.get_text()
  news_title_list.append(news_title_txt)
  
# print(news_title_list)  

# 2차 가공
'''
for i, title in enumerate(news_title_list):
  no = i + 1
  print(f"{no} : {title}")
'''  
'''
# 특정 키워드로 검색 후, 검색 결과 리스트 담고 출력
find_keyword_news_list = []  

keyword = '미국'
for i, title in enumerate(news_title_list):
  # no = i + 1
  # print(f"{no} : {title}")
  
  if title.find(keyword) != -1:
    find_keyword_news_list.append(title)
  
print(find_keyword_news_list) # 특정 키워드로 걸러진 데이터가 리스트에 저장

# 걸러진 데이터를 반복문을 이용해 출력
for i, title in enumerate(find_keyword_news_list):
  print(f"{i + 1} : {title} ")
'''  

print("== 검색 내용 출력 ==")
for i, title in enumerate(news_title_list):
  no = i + 1
  print(f"{no} : {title}")
  
print("== 데이터 저장 시작 ==")  

data = {
  '번호': range(1, len(news_title_list) + 1),
  '헤드 라인 뉴스 제목': news_title_list
  }

# 추출한 데이터를 엑셀에 저장
df = pd.DataFrame(data)
save_path = "C:\work\python_projects\뉴스_기사.xlsx"
df.to_excel(save_path, index=False)

# 엑셀 파일 열기
excel_file = 'C:\work\python_projects\뉴스_기사.xlsx'  # 엑셀 파일 경로
wb = load_workbook(excel_file)
ws = wb.active  # 현재 활성화된 시트 선택

# 1. **열 크기 조정**
# 'A'열(뉴스 제목 열)의 너비를 20으로 설정
ws.column_dimensions['A'].width = 15

# 'B'열(뉴스 링크 열)의 너비를 70으로 설정
ws.column_dimensions['B'].width = 100

# 2. **행 크기 조정**
# 1번 행(헤더 행)의 높이를 20로 설정
ws.row_dimensions[1].height = 25

# 가운데 정렬 스타일 정의
center_style = Alignment(horizontal='center', vertical='center')

# A열의 모든 셀을 가운데 정렬
for row in ws.iter_rows(min_col=1, max_col=1, min_row=1, max_row=ws.max_row):
    for cell in row:
        cell.alignment = center_style  # 셀의 정렬을 가운데로 설정

# B열의 1번 셀(B1)만 가운데 정렬
ws['B1'].alignment = center_style

wb.save(excel_file) # 새로운 파일로 저장

'''
# 3. **모든 열과 행의 크기를 조정**
# 특정 범위의 열과 행을 조정할 수도 있습니다.
# 예를 들어 A, B 열의 너비를 동시에 조정하는 경우
for col in ['A', 'B']:
    ws.column_dimensions[col].width = 20
'''

print("== 데이터 저장이 완료되었습니다. ==")

